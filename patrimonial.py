import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import json
import os
import csv
import math

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

FICHIER = "donnees.json"
CONFIG_FICHIER = "config.json"

# ─────────────────────────────────────────────
#  PALETTE VS CODE DARK
# ─────────────────────────────────────────────
BG_BASE      = "#1e1e1e"
BG_PANEL     = "#252526"
BG_WIDGET    = "#2d2d2d"
BG_HEADER    = "#333333"
BG_ROW_ODD   = "#1e1e1e"
BG_ROW_EVEN  = "#252526"
BG_SELECTED  = "#094771"
FG_PRIMARY   = "#d4d4d4"
FG_DIM       = "#858585"
FG_WHITE     = "#ffffff"
ACCENT       = "#007acc"
ACCENT_LIGHT = "#4fc1ff"
GREEN        = "#4ec9b0"
RED          = "#f44747"
YELLOW       = "#dcdcaa"
ORANGE       = "#ce9178"
FONT_MONO    = ("Consolas", 11)
FONT_BOLD    = ("Consolas", 11, "bold")
FONT_TITLE   = ("Consolas", 14, "bold")
FONT_SMALL   = ("Consolas", 9)

# ─────────────────────────────────────────────
#  STATE
# ─────────────────────────────────────────────
colonnes_revenu = ["Binance", "IBKR", "Tiktok", "Compte Bancaire", "Liquide", "Vinted"]

def charger_config():
    global colonnes_revenu
    if os.path.exists(CONFIG_FICHIER):
        try:
            with open(CONFIG_FICHIER, "r") as f:
                cfg = json.load(f)
                colonnes_revenu = cfg.get("colonnes", colonnes_revenu)
        except:
            pass

def sauvegarder_config():
    with open(CONFIG_FICHIER, "w") as f:
        json.dump({"colonnes": colonnes_revenu}, f)

# ─────────────────────────────────────────────
#  FENÊTRE PRINCIPALE
# ─────────────────────────────────────────────
charger_config()

fenetre = tk.Tk()
fenetre.title("Patrimonial")
fenetre.geometry("1400x780")
fenetre.configure(bg=BG_BASE)
fenetre.minsize(900, 600)

# ─────────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────────
header = tk.Frame(fenetre, bg=BG_PANEL, height=56)
header.pack(fill="x", side="top")
header.pack_propagate(False)

title_label = tk.Label(
    header, text="◈  PATRIMONIAL", font=("Consolas", 16, "bold"),
    bg=BG_PANEL, fg=ACCENT_LIGHT
)
title_label.pack(side="left", padx=20, pady=12)

label_dernier_total = tk.Label(
    header, text="Portfolio: — €", font=FONT_TITLE,
    bg=BG_PANEL, fg=FG_WHITE
)
label_dernier_total.pack(side="left", padx=30)

label_variation = tk.Label(
    header, text="", font=FONT_BOLD, bg=BG_PANEL, fg=FG_DIM
)
label_variation.pack(side="left", padx=5)

# ─────────────────────────────────────────────
#  TOOLBAR
# ─────────────────────────────────────────────
toolbar = tk.Frame(fenetre, bg=BG_BASE, height=40)
toolbar.pack(fill="x", padx=16, pady=(8, 0))

def btn_style(parent, text, command, color=ACCENT):
    b = tk.Button(
        parent, text=text, command=command,
        bg=BG_WIDGET, fg=FG_PRIMARY,
        font=FONT_SMALL, bd=0, padx=10, pady=5,
        relief="flat", cursor="hand2",
        activebackground=color, activeforeground=FG_WHITE
    )
    b.bind("<Enter>", lambda e: b.config(bg=color, fg=FG_WHITE))
    b.bind("<Leave>", lambda e: b.config(bg=BG_WIDGET, fg=FG_PRIMARY))
    return b

btn_add    = btn_style(toolbar, "＋  Ajouter ligne",     lambda: ajouter_ligne())
btn_del    = btn_style(toolbar, "－  Supprimer",         lambda: supprimer_ligne(), RED)
btn_cols   = btn_style(toolbar, "⚙  Colonnes",           lambda: ouvrir_gestion_colonnes(), YELLOW)
btn_graph  = btn_style(toolbar, "📈  Graphique",          lambda: ouvrir_graphique(), GREEN)
btn_csv    = btn_style(toolbar, "⬇  Export CSV",         lambda: exporter_csv())
btn_excel  = btn_style(toolbar, "⬇  Export Excel",       lambda: exporter_excel())

for b in [btn_add, btn_del, btn_cols, btn_graph, btn_csv, btn_excel]:
    b.pack(side="left", padx=3)

# ─────────────────────────────────────────────
#  TREEVIEW
# ─────────────────────────────────────────────
tree_frame = tk.Frame(fenetre, bg=BG_BASE)
tree_frame.pack(expand=True, fill="both", padx=16, pady=10)

scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical")
scrollbar_y.pack(side="right", fill="y")
scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal")
scrollbar_x.pack(side="bottom", fill="x")

style = ttk.Style()
style.theme_use("clam")
style.configure("Patrimonial.Treeview",
    background=BG_ROW_ODD,
    foreground=FG_PRIMARY,
    fieldbackground=BG_ROW_ODD,
    rowheight=30,
    font=FONT_MONO,
    borderwidth=0
)
style.configure("Patrimonial.Treeview.Heading",
    font=("Consolas", 10, "bold"),
    background=BG_HEADER,
    foreground=ACCENT_LIGHT,
    relief="flat",
    padding=(6, 6)
)
style.map("Patrimonial.Treeview",
    background=[("selected", BG_SELECTED)],
    foreground=[("selected", FG_WHITE)]
)
style.configure("Patrimonial.Vertical.TScrollbar",
    background=BG_WIDGET, troughcolor=BG_BASE,
    arrowcolor=FG_DIM, bordercolor=BG_BASE, relief="flat"
)
style.configure("Patrimonial.Horizontal.TScrollbar",
    background=BG_WIDGET, troughcolor=BG_BASE,
    arrowcolor=FG_DIM, bordercolor=BG_BASE, relief="flat"
)
scrollbar_y.configure(style="Patrimonial.Vertical.TScrollbar")
scrollbar_x.configure(style="Patrimonial.Horizontal.TScrollbar")

tree = None  # sera initialisé dans rebuild_tree

# ─────────────────────────────────────────────
#  CONSTRUCTION DYNAMIQUE DU TREEVIEW
# ─────────────────────────────────────────────
def get_all_colonnes():
    return ["Mois"] + colonnes_revenu + ["Total", "% Gain/Perte", "€ Gain/Perte"]

def rebuild_tree(donnees_existantes=None):
    global tree
    if tree is not None:
        tree.destroy()

    all_cols = get_all_colonnes()
    tree = ttk.Treeview(
        tree_frame, columns=all_cols, show="headings",
        height=20, style="Patrimonial.Treeview",
        yscrollcommand=scrollbar_y.set,
        xscrollcommand=scrollbar_x.set
    )
    scrollbar_y.config(command=tree.yview)
    scrollbar_x.config(command=tree.xview)

    for col in all_cols:
        tree.heading(col, text=col)
        if col == "Mois":
            tree.column(col, width=100, anchor="center", minwidth=80)
        elif col in ("Total", "€ Gain/Perte"):
            tree.column(col, width=120, anchor="e", minwidth=90)
        elif col == "% Gain/Perte":
            tree.column(col, width=110, anchor="center", minwidth=90)
        else:
            tree.column(col, width=120, anchor="e", minwidth=80)

    tree.tag_configure("oddrow",  background=BG_ROW_ODD,  foreground=FG_PRIMARY)
    tree.tag_configure("evenrow", background=BG_ROW_EVEN, foreground=FG_PRIMARY)
    tree.tag_configure("positive", foreground=GREEN)
    tree.tag_configure("negative", foreground=RED)
    tree.tag_configure("neutral",  foreground=FG_DIM)

    tree.pack(expand=True, fill="both")
    tree.bind("<Double-1>", modifier_cellule)

    if donnees_existantes:
        n_rev = len(colonnes_revenu)
        for row in donnees_existantes:
            mois = row[0] if len(row) > 0 else "Janvier"
            revenus = list(row[1:1+n_rev]) if len(row) >= 1+n_rev else [0]*n_rev
            # Pad si nécessaire
            while len(revenus) < n_rev:
                revenus.append(0)
            revenus = revenus[:n_rev]
            tree.insert("", "end", values=[mois] + revenus + [0, "0 %", 0])

    recalculer_tout()

# ─────────────────────────────────────────────
#  LOGIQUE CALCUL
# ─────────────────────────────────────────────
def n_rev():
    return len(colonnes_revenu)

def convertir_nombre(valeur):
    valeur = str(valeur).replace(",", ".").strip()
    try:
        return float(valeur)
    except:
        return None

def recalculer_total_ligne(item):
    values = tree.item(item)["values"]
    try:
        total = sum(float(values[i]) for i in range(1, 1 + n_rev()))
    except:
        total = 0
    tree.set(item, "Total", round(total, 2))

def recalculer_pourcentage():
    items = tree.get_children()
    for i, item in enumerate(items):
        if i == 0:
            tree.set(item, "% Gain/Perte", "—")
            tree.set(item, "€ Gain/Perte", "—")
        else:
            try:
                t_act  = float(tree.set(item,       "Total"))
                t_prec = float(tree.set(items[i-1], "Total"))
            except:
                t_act = t_prec = 0
            if t_prec == 0:
                pct = 0
            else:
                pct = ((t_act - t_prec) / t_prec) * 100
            gain = t_act - t_prec
            tree.set(item, "% Gain/Perte", f"{round(pct, 2):+.2f} %")
            tree.set(item, "€ Gain/Perte", f"{round(gain, 2):+.2f}")

def recalculer_tout():
    if tree is None:
        return
    for item in tree.get_children():
        recalculer_total_ligne(item)
    recalculer_pourcentage()
    appliquer_tags()
    mettre_a_jour_header()

def appliquer_tags():
    items = tree.get_children()
    for i, item in enumerate(items):
        base_tag = "evenrow" if i % 2 == 0 else "oddrow"
        try:
            gain = tree.set(item, "€ Gain/Perte")
            val = float(gain.replace("+", ""))
            if val > 0:
                color_tag = "positive"
            elif val < 0:
                color_tag = "negative"
            else:
                color_tag = base_tag
        except:
            color_tag = base_tag
        tree.item(item, tags=(base_tag, color_tag))

def mettre_a_jour_header():
    items = tree.get_children()
    dernier_total = 0
    dernier_mois  = ""
    variation_str = ""

    for item in reversed(items):
        try:
            t = float(tree.set(item, "Total"))
            if t > 0:
                dernier_total = t
                dernier_mois  = tree.set(item, "Mois")
                gain = tree.set(item, "€ Gain/Perte")
                pct  = tree.set(item, "% Gain/Perte")
                variation_str = f"{gain} €  ({pct})"
                try:
                    g = float(gain.replace("+",""))
                    label_variation.config(fg=GREEN if g > 0 else (RED if g < 0 else FG_DIM))
                except:
                    label_variation.config(fg=FG_DIM)
                break
        except:
            pass

    label_dernier_total.config(
        text=f"Portfolio ({dernier_mois}): {dernier_total:,.2f} €" if dernier_mois else "Portfolio: — €"
    )
    label_variation.config(text=variation_str)

# ─────────────────────────────────────────────
#  MOIS PRÉDÉFINIS
# ─────────────────────────────────────────────
MOIS_LIST = [
    "Janvier","Février","Mars","Avril","Mai","Juin",
    "Juillet","Août","Septembre","Octobre","Novembre","Décembre"
]

def prochain_mois():
    items = tree.get_children()
    if not items:
        return MOIS_LIST[0]
    dernier = tree.set(items[-1], "Mois")
    if dernier in MOIS_LIST:
        idx = MOIS_LIST.index(dernier)
        return MOIS_LIST[(idx + 1) % 12]
    return MOIS_LIST[0]

def ajouter_ligne():
    mois = prochain_mois()
    vals = [mois] + [0]*n_rev() + [0, "—", "—"]
    tree.insert("", "end", values=vals)
    recalculer_tout()
    sauvegarder_donnees()

def supprimer_ligne():
    for item in tree.selection():
        tree.delete(item)
    recalculer_tout()
    sauvegarder_donnees()

# ─────────────────────────────────────────────
#  ÉDITION CELLULE
# ─────────────────────────────────────────────
def modifier_cellule(event):
    item = tree.identify_row(event.y)
    col  = tree.identify_column(event.x)
    if not item or not col:
        return
    col_idx = int(col.replace("#", "")) - 1
    all_cols = get_all_colonnes()
    if col_idx >= len(all_cols):
        return
    col_name = all_cols[col_idx]
    # Colonnes calculées → non éditables
    if col_name in ("Total", "% Gain/Perte", "€ Gain/Perte"):
        return

    try:
        bbox = tree.bbox(item, col)
        if not bbox:
            return
        x, y, w, h = bbox
    except:
        return

    if col_name == "Mois":
        # Dropdown mois
        var = tk.StringVar(value=tree.set(item, col_name))
        combo = ttk.Combobox(tree, textvariable=var, values=MOIS_LIST,
                             state="readonly", font=FONT_MONO)
        style.configure("TCombobox",
            fieldbackground=BG_WIDGET, background=BG_WIDGET,
            foreground=FG_WHITE, selectbackground=BG_SELECTED)
        combo.place(x=x, y=y, width=w, height=h)
        combo.focus()

        def on_select(e=None):
            tree.set(item, col_name, var.get())
            combo.destroy()
            sauvegarder_donnees()

        combo.bind("<<ComboboxSelected>>", on_select)
        combo.bind("<FocusOut>", lambda e: combo.destroy())
    else:
        entry = tk.Entry(tree, font=FONT_MONO, bg=BG_WIDGET, fg=FG_WHITE,
                         insertbackground=ACCENT_LIGHT, bd=0,
                         highlightthickness=1, highlightcolor=ACCENT)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, tree.item(item)["values"][col_idx])
        entry.select_range(0, "end")
        entry.focus()

        def sauvegarder(e=None):
            val = entry.get()
            nb = convertir_nombre(val)
            if nb is None:
                entry.destroy()
                return
            tree.set(item, col_name, round(nb, 2))
            entry.destroy()
            recalculer_tout()
            sauvegarder_donnees()

        entry.bind("<Return>", sauvegarder)
        entry.bind("<FocusOut>", lambda e: entry.destroy())

# ─────────────────────────────────────────────
#  GESTION COLONNES (modale)
# ─────────────────────────────────────────────
def ouvrir_gestion_colonnes():
    win = tk.Toplevel(fenetre)
    win.title("Gérer les colonnes")
    win.configure(bg=BG_BASE)
    win.geometry("420x460")
    win.resizable(False, False)
    win.grab_set()

    tk.Label(win, text="⚙  Colonnes de revenus", font=("Consolas", 13, "bold"),
             bg=BG_BASE, fg=ACCENT_LIGHT).pack(pady=(16,8))

    sep = tk.Frame(win, bg=ACCENT, height=1)
    sep.pack(fill="x", padx=20)

    list_frame = tk.Frame(win, bg=BG_PANEL, bd=0)
    list_frame.pack(fill="both", expand=True, padx=20, pady=10)

    listbox = tk.Listbox(list_frame, bg=BG_WIDGET, fg=FG_PRIMARY,
                         font=FONT_MONO, selectbackground=BG_SELECTED,
                         selectforeground=FG_WHITE, bd=0, highlightthickness=0,
                         activestyle="none")
    listbox.pack(fill="both", expand=True, padx=6, pady=6)
    for c in colonnes_revenu:
        listbox.insert("end", c)

    ctrl = tk.Frame(win, bg=BG_BASE)
    ctrl.pack(fill="x", padx=20, pady=(0,8))

    entry_new = tk.Entry(ctrl, font=FONT_MONO, bg=BG_WIDGET, fg=FG_WHITE,
                         insertbackground=ACCENT_LIGHT, bd=0,
                         highlightthickness=1, highlightcolor=ACCENT)
    entry_new.pack(side="left", fill="x", expand=True, padx=(0,6), ipady=4)
    entry_new.insert(0, "Nouvelle colonne…")
    entry_new.bind("<FocusIn>",  lambda e: entry_new.delete(0,"end") if entry_new.get()=="Nouvelle colonne…" else None)
    entry_new.bind("<FocusOut>", lambda e: entry_new.insert(0,"Nouvelle colonne…") if entry_new.get()=="" else None)

    def ajouter_col():
        nom = entry_new.get().strip()
        if nom and nom != "Nouvelle colonne…" and nom not in colonnes_revenu:
            colonnes_revenu.append(nom)
            listbox.insert("end", nom)
            entry_new.delete(0,"end")
            entry_new.insert(0,"Nouvelle colonne…")

    def supprimer_col():
        sel = listbox.curselection()
        if not sel or len(colonnes_revenu) <= 1:
            return
        idx = sel[0]
        colonnes_revenu.pop(idx)
        listbox.delete(idx)

    def renommer_col():
        sel = listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        nouveau = simpledialog.askstring("Renommer", f"Nouveau nom pour '{colonnes_revenu[idx]}' :",
                                          initialvalue=colonnes_revenu[idx], parent=win)
        if nouveau and nouveau.strip():
            colonnes_revenu[idx] = nouveau.strip()
            listbox.delete(idx)
            listbox.insert(idx, nouveau.strip())

    def monter():
        sel = listbox.curselection()
        if not sel or sel[0] == 0: return
        i = sel[0]
        colonnes_revenu[i-1], colonnes_revenu[i] = colonnes_revenu[i], colonnes_revenu[i-1]
        txt = listbox.get(i)
        listbox.delete(i)
        listbox.insert(i-1, txt)
        listbox.selection_set(i-1)

    def descendre():
        sel = listbox.curselection()
        if not sel or sel[0] >= listbox.size()-1: return
        i = sel[0]
        colonnes_revenu[i], colonnes_revenu[i+1] = colonnes_revenu[i+1], colonnes_revenu[i]
        txt = listbox.get(i)
        listbox.delete(i)
        listbox.insert(i+1, txt)
        listbox.selection_set(i+1)

    btn_frame = tk.Frame(win, bg=BG_BASE)
    btn_frame.pack(fill="x", padx=20, pady=(0,4))

    for text, cmd, col in [
        ("＋ Ajouter", ajouter_col, ACCENT),
        ("✎ Renommer", renommer_col, YELLOW),
        ("✕ Supprimer", supprimer_col, RED),
        ("↑", monter, FG_DIM),
        ("↓", descendre, FG_DIM),
    ]:
        b = tk.Button(btn_frame, text=text, command=cmd,
                      bg=BG_WIDGET, fg=FG_PRIMARY, font=FONT_SMALL,
                      bd=0, padx=8, pady=4, relief="flat", cursor="hand2")
        b.bind("<Enter>", lambda e, b=b, c=col: b.config(bg=c, fg=FG_WHITE))
        b.bind("<Leave>", lambda e, b=b: b.config(bg=BG_WIDGET, fg=FG_PRIMARY))
        b.pack(side="left", padx=3)

    def appliquer():
        donnees = extraire_donnees_brutes()
        sauvegarder_config()
        rebuild_tree(donnees)
        sauvegarder_donnees()
        win.destroy()

    tk.Button(win, text="✓  Appliquer", command=appliquer,
              bg=ACCENT, fg=FG_WHITE, font=FONT_BOLD,
              bd=0, padx=16, pady=8, relief="flat", cursor="hand2"
    ).pack(pady=(4,14))

def extraire_donnees_brutes():
    rows = []
    for item in tree.get_children():
        vals = tree.item(item)["values"]
        rows.append(vals)
    return rows

# ─────────────────────────────────────────────
#  SAUVEGARDE / CHARGEMENT
# ─────────────────────────────────────────────
def sauvegarder_donnees():
    if tree is None:
        return
    data = []
    for item in tree.get_children():
        vals = list(tree.item(item)["values"])
        # Ne sauvegarder que Mois + revenus
        data.append(vals[:1 + n_rev()])
    with open(FICHIER, "w") as f:
        json.dump(data, f, indent=2)

def charger_donnees():
    if not os.path.exists(FICHIER):
        return []
    try:
        with open(FICHIER, "r") as f:
            return json.load(f)
    except:
        return []

# ─────────────────────────────────────────────
#  GRAPHIQUE
# ─────────────────────────────────────────────
def ouvrir_graphique():
    items = tree.get_children()
    if not items:
        messagebox.showinfo("Graphique", "Aucune donnée à afficher.")
        return

    labels = []
    totaux = []
    gains  = []
    for item in items:
        try:
            t = float(tree.set(item, "Total"))
        except:
            t = 0
        labels.append(tree.set(item, "Mois"))
        totaux.append(t)

    for i in range(len(totaux)):
        if i == 0:
            gains.append(0)
        else:
            gains.append(totaux[i] - totaux[i-1])

    win = tk.Toplevel(fenetre)
    win.title("Évolution du patrimoine")
    win.configure(bg=BG_BASE)
    win.geometry("900x500")
    win.resizable(True, True)

    canvas = tk.Canvas(win, bg=BG_BASE, bd=0, highlightthickness=0)
    canvas.pack(fill="both", expand=True, padx=10, pady=10)

    def dessiner(event=None):
        canvas.delete("all")
        W = canvas.winfo_width()
        H = canvas.winfo_height()
        if W < 100 or H < 100:
            return

        pad_l, pad_r, pad_t, pad_b = 70, 30, 40, 60
        gw = W - pad_l - pad_r
        gh = H - pad_t - pad_b

        # Titre
        canvas.create_text(W//2, 18, text="Évolution du Patrimoine",
                            font=("Consolas", 13, "bold"), fill=ACCENT_LIGHT)

        if not totaux or max(totaux) == 0:
            canvas.create_text(W//2, H//2, text="Pas de données",
                               font=FONT_MONO, fill=FG_DIM)
            return

        max_v = max(totaux) * 1.1 if max(totaux) > 0 else 1
        min_v = min(0, min(totaux))
        rng   = max_v - min_v if max_v != min_v else 1

        def x_pos(i):
            n = len(labels)
            if n <= 1:
                return pad_l + gw // 2
            return pad_l + int(i * gw / (n - 1))

        def y_pos(v):
            return pad_t + gh - int((v - min_v) / rng * gh)

        # Grille
        for k in range(6):
            val  = min_v + rng * k / 5
            yy   = y_pos(val)
            canvas.create_line(pad_l, yy, pad_l + gw, yy,
                               fill="#333333", dash=(4, 4))
            canvas.create_text(pad_l - 6, yy,
                               text=f"{val:,.0f}€", anchor="e",
                               font=("Consolas", 8), fill=FG_DIM)

        # Axe X
        canvas.create_line(pad_l, pad_t + gh, pad_l + gw, pad_t + gh,
                           fill="#555555", width=1)
        # Axe Y
        canvas.create_line(pad_l, pad_t, pad_l, pad_t + gh,
                           fill="#555555", width=1)

        # Barres gain/perte (fond)
        bar_w = max(4, gw // (len(labels) * 2))
        for i, g in enumerate(gains):
            if g == 0:
                continue
            xc = x_pos(i)
            y0 = y_pos(0)
            yg = y_pos(g) if g > 0 else y_pos(0)
            yb = y_pos(0) if g > 0 else y_pos(g)
            col = "#1a4a2a" if g > 0 else "#4a1a1a"
            canvas.create_rectangle(xc - bar_w//2, yg, xc + bar_w//2, yb,
                                     fill=col, outline="")

        # Ligne principale (dégradé simulé par segments)
        pts = [(x_pos(i), y_pos(v)) for i, v in enumerate(totaux)]
        if len(pts) >= 2:
            for i in range(len(pts) - 1):
                t_ratio = i / max(1, len(pts) - 2)
                # couleur interpolée bleu→cyan
                r = int(0   + t_ratio * 79)
                g_c = int(122 + t_ratio * 77)
                b_c = int(204 + t_ratio * 51)
                col = f"#{r:02x}{g_c:02x}{b_c:02x}"
                canvas.create_line(pts[i][0], pts[i][1],
                                   pts[i+1][0], pts[i+1][1],
                                   fill=col, width=3, capstyle="round")

        # Points & labels
        for i, (px, py) in enumerate(pts):
            # Halo
            canvas.create_oval(px-7, py-7, px+7, py+7,
                               fill=BG_BASE, outline=ACCENT, width=2)
            canvas.create_oval(px-3, py-3, px+3, py+3,
                               fill=ACCENT_LIGHT, outline="")
            # Valeur
            canvas.create_text(px, py - 16, text=f"{totaux[i]:,.0f}€",
                               font=("Consolas", 8, "bold"), fill=FG_WHITE)
            # Label mois
            canvas.create_text(px, pad_t + gh + 14, text=labels[i],
                               font=("Consolas", 8), fill=FG_DIM, angle=0)

    canvas.bind("<Configure>", dessiner)
    win.after(100, dessiner)

# ─────────────────────────────────────────────
#  EXPORTS
# ─────────────────────────────────────────────
def get_header():
    return ["Mois"] + colonnes_revenu + ["Total", "% Gain/Perte", "€ Gain/Perte"]

def exporter_csv():
    from tkinter import filedialog
    path = filedialog.asksaveasfilename(
        defaultextension=".csv", filetypes=[("CSV", "*.csv")],
        initialfile="patrimonial_export.csv"
    )
    if not path:
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(get_header())
        for item in tree.get_children():
            writer.writerow(tree.item(item)["values"])
    messagebox.showinfo("Export CSV", f"Fichier exporté :\n{path}")

def exporter_excel():
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Excel indisponible",
            "openpyxl n'est pas installé.\nLancez : pip install openpyxl")
        return
    from tkinter import filedialog
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
        initialfile="patrimonial_export.xlsx"
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patrimoine"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header
    header = get_header()
    for ci, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font  = Font(bold=True, color="4FC1FF", name="Consolas")
        cell.fill  = PatternFill("solid", fgColor="333333")
        cell.alignment = Alignment(horizontal="center")

    thin = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, item in enumerate(tree.get_children(), 2):
        vals = tree.item(item)["values"]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Consolas", color="D4D4D4")
            cell.fill      = PatternFill("solid", fgColor="252526" if ri%2==0 else "1E1E1E")
            cell.alignment = Alignment(horizontal="right" if ci > 1 else "left")
            cell.border    = border

    for ci in range(1, len(header)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    wb.save(path)
    messagebox.showinfo("Export Excel", f"Fichier exporté :\n{path}")

# ─────────────────────────────────────────────
#  STATUS BAR
# ─────────────────────────────────────────────
status_bar = tk.Frame(fenetre, bg=ACCENT, height=22)
status_bar.pack(fill="x", side="bottom")
status_bar.pack_propagate(False)

status_label = tk.Label(status_bar, text="  Double-clic pour éditer une cellule  •  Entrée pour valider",
                         font=("Consolas", 9), bg=ACCENT, fg=FG_WHITE, anchor="w")
status_label.pack(side="left", fill="x", padx=8)

version_label = tk.Label(status_bar, text="Patrimonial v2.0  ",
                          font=("Consolas", 9), bg=ACCENT, fg=FG_WHITE)
version_label.pack(side="right")

# ─────────────────────────────────────────────
#  RACCOURCIS CLAVIER
# ─────────────────────────────────────────────
fenetre.bind("<Control-n>", lambda e: ajouter_ligne())
fenetre.bind("<Delete>",    lambda e: supprimer_ligne())
fenetre.bind("<Control-e>", lambda e: exporter_csv())
fenetre.bind("<Control-g>", lambda e: ouvrir_graphique())

# ─────────────────────────────────────────────
#  INIT
# ─────────────────────────────────────────────
donnees = charger_donnees()
rebuild_tree(donnees)

fenetre.mainloop()
